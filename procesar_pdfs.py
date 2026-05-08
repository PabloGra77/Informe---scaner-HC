import argparse
import functools
from concurrent.futures import FIRST_COMPLETED, ProcessPoolExecutor, wait
from datetime import datetime
import os
import re
import shutil
import sys
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Tuple

# Asegura que cada print llegue inmediatamente al proceso padre (la GUI),
# evitando el buffering por bloque cuando stdout esta conectado a un PIPE.
print = functools.partial(print, flush=True)  # type: ignore[assignment]
try:
    sys.stdout.reconfigure(line_buffering=True)  # type: ignore[attr-defined]
except Exception:
    pass

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader


# Patrones para extraer los campos marcados en rojo.
FIELD_PATTERNS: Dict[str, List[str]] = {
    "codigo_actividad": [
        r"(?is)\n\s*([^\n\r]{3,160}?\s*\-\s*[^\n\r]{3,160})\s*[\r\n]+\s*IDENTIFICACI[OÓ]N\s+DEL\s+PACIENTE\b",
        r"(?im)^\s*([A-Z0-9]{3,20}\s*\-\s*[^\n\r]{3,160})\s*$",
    ],
    "programa": [
        r"(?is)\n\s*(?!Copia\s*$)([^\n\r]{4,160})\s*[\r\n]+\s*Apellidos\s*:",
        r"(?is)\bCENTRO\s+DE\s+REHABILITACI[OÓ]N\s+INTEGRAL\b[^\n\r]*[\r\n]+\s*[^\n\r]*[\r\n]+\s*(?!Copia\s*$)([^\n\r]{4,160})",
    ],
    "apellidos": [
        r"(?im)\bApellidos\s*:\s*([^\n\r:]{2,80}?)(?=\s+Nombres\s*:)",
    ],
    "nombres": [
        r"(?im)\bNombres\s*:\s*([^\n\r:]{2,80})",
    ],
    "tipo_identificacion": [
        r"(?im)\bTipo\s+Identificaci[oó]n\s*:\s*([A-Z0-9]{1,10})",
    ],
    "numero_documento": [
        r"(?im)\bN[uú]mero\s+documento\s*:\s*([A-Z0-9][A-Z0-9\-\.]{2,})",
        r"(?im)\b(?:n[uú]mero|no\.?|nro\.?)\s*de\s*documento\b\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-\.]{2,})",
        r"(?im)\b(?:documento|doc(?:umento)?)\b\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-\.]{2,})",
        r"(?im)\b(?:identificaci[oó]n|id)\b\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-\.]{2,})",
    ],
}

HEADERS = [
    "archivo_original",
    "archivo_renombrado",
    "ruta_archivo",
    "codigo_actividad",
    "programa",
    "apellidos",
    "nombres",
    "tipo_identificacion",
    "numero_documento",
    "fecha_atencion",
    "novedades",
    "fecha_novedad",
    "estado",
    "observacion",
]

HEADERS_ENCARPETADO = [
    "ruta_archivo",
    "numero_documento",
    "carpeta_destino",
    "archivo_destino",
    "estado",
    "observacion",
]

HEADERS_NUEVOS = [
    "ruta_archivo",
    "archivo",
    "paciente",
    "numero_ide",
    "ingreso",
    "tipo_servicio",
    "fecha_atencion",
    "diagnostico_tipo",
    "diagnostico_clase",
    "diagnostico_codigo",
    "diagnostico_descripcion",
    "nota",
    "novedades",
    "fecha_novedad",
    "estado",
    "observacion",
]


def extract_pdf_text(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    texts: List[str] = []
    for page in reader.pages:
        texts.append(page.extract_text() or "")
    return "\n".join(texts)


def clean_value(value: str) -> str:
    value = re.sub(r"\s+", " ", value).strip(" .:-\t\r\n")
    return value


def extract_fecha_atencion(text: str) -> str:
    patterns = [
        r"(?is)\bCONSULTA\b.*?\bFecha\s*(?:de\s*)?atenci[oó]n\s*:?\s*([0-3]?\d[\/-][0-1]?\d[\/-](?:\d{4}|\d{2})(?:\s+[0-2]?\d:\d{2}(?::\d{2})?)?)",
        r"(?is)\bFecha\s*(?:de\s*)?atenci[oó]n\s*:?\s*([0-3]?\d[\/-][0-1]?\d[\/-](?:\d{4}|\d{2})(?:\s+[0-2]?\d:\d{2}(?::\d{2})?)?)",
        r"(?is)\bFechaatenci[oó]n\s*:?\s*([0-3]?\d[\/-][0-1]?\d[\/-](?:\d{4}|\d{2})(?:\s+[0-2]?\d:\d{2}(?::\d{2})?)?)",
        r"(?im)\bFecha\s*(?:de\s*)?\s*:?\s*([0-3]?\d[\/-][0-1]?\d[\/-](?:\d{4}|\d{2}))",
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return clean_value(match.group(1))

    return ""


def detect_fields(text: str) -> Dict[str, str]:
    found: Dict[str, str] = {}
    for field, patterns in FIELD_PATTERNS.items():
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                found[field] = clean_value(match.group(1))
                break
        if field not in found:
            found[field] = ""

    # Reintentos robustos si se capturan valores no validos por ruido de OCR/plantilla.
    if found.get("programa", "").strip().upper() == "COPIA":
        fallback_programa = re.search(
            r"(?is)\n\s*(?!Copia\s*$)([^\n\r]{4,160})\s*[\r\n]+\s*Apellidos\s*:",
            text,
        )
        if fallback_programa:
            found["programa"] = clean_value(fallback_programa.group(1))

    if not found.get("codigo_actividad", ""):
        fallback_codigo = re.search(
            r"(?is)\n\s*([^\n\r]{3,160}?\s*\-\s*[^\n\r]{3,160})\s*[\r\n]+\s*IDENTIFICACI[OÓ]N\s+DEL\s+PACIENTE\b",
            text,
        )
        if fallback_codigo:
            found["codigo_actividad"] = clean_value(fallback_codigo.group(1))

    found["fecha_atencion"] = extract_fecha_atencion(text)

    novedades, fecha_novedad = _extract_novedades(text)
    found["novedades"] = novedades
    found["fecha_novedad"] = fecha_novedad

    return found


def sanitize_filename(value: str) -> str:
    value = value.strip().upper()
    value = re.sub(r"[^A-Z0-9\-_.]", "_", value)
    value = re.sub(r"_+", "_", value)
    value = value.strip("._-")
    return value


def date_to_ddmmyyyy(value: str) -> str:
    match = re.search(r"([0-3]?\d)[\/-]([0-1]?\d)[\/-]((?:\d{4}|\d{2}))", value or "")
    if not match:
        return ""

    day = int(match.group(1))
    month = int(match.group(2))
    year = int(match.group(3))
    if year < 100:
        year += 2000

    try:
        dt = datetime(year, month, day)
    except ValueError:
        return ""

    return dt.strftime("%d%m%Y")


def build_renamed_stem(numero_documento: str, fecha_atencion: str) -> str:
    doc = sanitize_filename(numero_documento)
    fecha = date_to_ddmmyyyy(fecha_atencion)

    if doc and fecha:
        return f"{doc}_{fecha}"
    return doc


def unique_target_path(original_path: Path, desired_stem: str) -> Path:
    candidate = original_path.with_name(f"{desired_stem}.pdf")
    if not candidate.exists() or candidate.samefile(original_path):
        return candidate

    i = 2
    while True:
        candidate = original_path.with_name(f"{desired_stem}_{i}.pdf")
        if not candidate.exists():
            return candidate
        i += 1


def unique_target_file_path(folder: Path, file_name: str) -> Path:
    candidate = folder / file_name
    if not candidate.exists():
        return candidate

    stem = Path(file_name).stem
    suffix = Path(file_name).suffix
    i = 2
    while True:
        candidate = folder / f"{stem}_{i}{suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def extract_row_data(pdf_path: Path) -> Dict[str, str]:
    row: Dict[str, str] = {
        "_source_path": str(pdf_path.resolve()),
        "_desired_stem": "",
        "archivo_original": pdf_path.name,
        "archivo_renombrado": "",
        "ruta_archivo": str(pdf_path.resolve()),
        "estado": "OK",
        "observacion": "",
    }

    try:
        text = extract_pdf_text(pdf_path)
        fields = detect_fields(text)
        row.update(fields)

        numero_documento = fields.get("numero_documento", "")
        if not numero_documento:
            row["estado"] = "SIN_NUMERO"
            row["observacion"] = "No se detecto numero_documento en el PDF"
            return row

        safe_stem = build_renamed_stem(numero_documento, fields.get("fecha_atencion", ""))
        if not safe_stem:
            row["estado"] = "SIN_NUMERO"
            row["observacion"] = "numero_documento detectado pero invalido para nombre de archivo"
            return row

        row["_desired_stem"] = safe_stem
        row["archivo_renombrado"] = f"{safe_stem}.pdf"
        if not fields.get("fecha_atencion", ""):
            row["observacion"] = "No se detecto fecha_atencion; se renombra solo con numero_documento"

        return row
    except Exception as exc:
        row["estado"] = "ERROR"
        row["observacion"] = str(exc)
        return row


class ExcelPartWriter:
    def __init__(
        self,
        base_output_path: Path,
        max_rows_per_file: int,
        headers: List[str] = None,
        sheet_name: str = "Informe",
    ) -> None:
        self.base_output_path = base_output_path
        self.max_rows_per_file = max_rows_per_file
        self.headers = headers if headers is not None else HEADERS
        self.sheet_name = sheet_name
        self.part_number = 0
        self.rows_in_part = 0
        self.generated_files: List[Path] = []
        self.wb = None
        self.ws = None
        self._start_new_part()

    def _part_path(self, part_number: int) -> Path:
        if part_number == 1:
            return self.base_output_path
        stem = self.base_output_path.stem
        suffix = self.base_output_path.suffix or ".xlsx"
        return self.base_output_path.with_name(f"{stem}_PARTE_{part_number:04d}{suffix}")

    def _start_new_part(self) -> None:
        self.part_number += 1
        self.rows_in_part = 0
        self.wb = Workbook(write_only=True)
        self.ws = self.wb.create_sheet(self.sheet_name)
        self.ws.append(self.headers)

    def append(self, row: Dict[str, str]) -> None:
        if self.rows_in_part >= self.max_rows_per_file:
            self.close_current_part()
            self._start_new_part()
        self.ws.append([row.get(h, "") for h in self.headers])
        self.rows_in_part += 1

    def close_current_part(self) -> None:
        if self.wb is None:
            return
        path = self._part_path(self.part_number)
        self.wb.save(path)
        self.generated_files.append(path)
        self.wb = None
        self.ws = None

    def close(self) -> List[Path]:
        self.close_current_part()
        return self.generated_files


def iter_pdf_files(input_path: Path, recursive: bool) -> Iterator[Path]:
    if recursive:
        stack = [input_path]
        while stack:
            current = stack.pop()
            try:
                with os.scandir(current) as entries:
                    for entry in entries:
                        try:
                            if entry.is_file(follow_symlinks=False) and entry.name.lower().endswith(".pdf"):
                                yield Path(entry.path)
                            elif entry.is_dir(follow_symlinks=False):
                                stack.append(Path(entry.path))
                        except PermissionError:
                            continue
            except PermissionError:
                continue
        return

    with os.scandir(input_path) as entries:
        for entry in entries:
            try:
                if entry.is_file(follow_symlinks=False) and entry.name.lower().endswith(".pdf"):
                    yield Path(entry.path)
            except PermissionError:
                continue


def count_pdf_files(input_path: Path, recursive: bool) -> int:
    count = 0
    for _ in iter_pdf_files(input_path, recursive):
        count += 1
    return count


def resolve_input_dir(raw_input: str) -> Path:
    candidate = Path(raw_input).expanduser().resolve()
    if candidate.exists() and candidate.is_dir():
        return candidate
    raise SystemExit(f"La ruta de entrada no es una carpeta valida: {candidate}")


def normalize_header(value: str) -> str:
    value = (value or "").strip().lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    value = re.sub(r"\s+", "_", value)
    return value


def resolve_file_path_from_excel(raw_path: str, excel_path: Path) -> Path:
    candidate = Path((raw_path or "").strip().strip('"')).expanduser()
    if candidate.is_absolute():
        return candidate.resolve()
    return (excel_path.parent / candidate).resolve()


def sanitize_folder_name(value: str) -> str:
    value = value.strip().upper()
    value = re.sub(r"[^A-Z0-9\-_.]", "_", value)
    value = re.sub(r"_+", "_", value)
    value = value.strip("._-")
    return value


def process_excel_foldering(
    excel_input: Path,
    excel_output: Path,
    ruta_column_name: str,
    documento_column_name: str,
) -> Tuple[int, int, int]:
    wb = load_workbook(excel_input, data_only=True, read_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise SystemExit("El Excel no tiene encabezados")

    header_map: Dict[str, int] = {}
    for idx, raw_header in enumerate(header_row):
        key = normalize_header(str(raw_header or ""))
        if key and key not in header_map:
            header_map[key] = idx

    ruta_key = normalize_header(ruta_column_name)
    doc_key = normalize_header(documento_column_name)

    if ruta_key not in header_map:
        raise SystemExit(f"No se encontro la columna '{ruta_column_name}' en el Excel")
    if doc_key not in header_map:
        raise SystemExit(f"No se encontro la columna '{documento_column_name}' en el Excel")

    ruta_idx = header_map[ruta_key]
    doc_idx = header_map[doc_key]

    output_wb = Workbook(write_only=True)
    output_ws = output_wb.create_sheet("Informe_Encarpetado")
    output_ws.append(HEADERS_ENCARPETADO)

    total = 0
    ok = 0
    errores = 0

    for row_values in rows:
        total += 1

        raw_ruta = ""
        raw_doc = ""
        if ruta_idx < len(row_values):
            raw_ruta = str(row_values[ruta_idx] or "").strip()
        if doc_idx < len(row_values):
            raw_doc = str(row_values[doc_idx] or "").strip()

        out_row = {
            "ruta_archivo": raw_ruta,
            "numero_documento": raw_doc,
            "carpeta_destino": "",
            "archivo_destino": "",
            "estado": "OK",
            "observacion": "",
        }

        if not raw_ruta:
            out_row["estado"] = "ERROR"
            out_row["observacion"] = "ruta_archivo vacia"
            errores += 1
            output_ws.append([out_row[h] for h in HEADERS_ENCARPETADO])
            continue

        if not raw_doc:
            out_row["estado"] = "ERROR"
            out_row["observacion"] = "numero_documento vacio"
            errores += 1
            output_ws.append([out_row[h] for h in HEADERS_ENCARPETADO])
            continue

        doc_folder_name = sanitize_folder_name(raw_doc)
        if not doc_folder_name:
            out_row["estado"] = "ERROR"
            out_row["observacion"] = "numero_documento invalido para carpeta"
            errores += 1
            output_ws.append([out_row[h] for h in HEADERS_ENCARPETADO])
            continue

        source = resolve_file_path_from_excel(raw_ruta, excel_input)
        if not source.exists() or not source.is_file():
            out_row["estado"] = "ERROR"
            out_row["observacion"] = "Archivo no encontrado"
            errores += 1
            output_ws.append([out_row[h] for h in HEADERS_ENCARPETADO])
            continue

        try:
            target_folder = source.parent / doc_folder_name
            target_folder.mkdir(parents=True, exist_ok=True)

            # Preferencia del usuario: nombres de salida en mayuscula.
            target_name = source.name.upper()
            target = unique_target_file_path(target_folder, target_name)
            shutil.move(str(source), str(target))

            out_row["carpeta_destino"] = str(target_folder.resolve())
            out_row["archivo_destino"] = str(target.resolve())
            ok += 1
        except Exception as exc:
            out_row["estado"] = "ERROR"
            out_row["observacion"] = f"Error al mover: {exc}"
            errores += 1

        output_ws.append([out_row[h] for h in HEADERS_ENCARPETADO])

    output_wb.save(excel_output)
    return total, ok, errores


def apply_rename(row: Dict[str, str], rename: bool) -> Dict[str, str]:
    source = Path(row.get("_source_path", ""))
    desired_stem = row.get("_desired_stem", "")

    row.pop("_source_path", None)
    row.pop("_desired_stem", None)

    if not source.exists():
        row["estado"] = "ERROR"
        row["observacion"] = "Archivo no encontrado al intentar renombrar"
        return row

    if row.get("estado") != "OK" or not desired_stem:
        row["archivo_renombrado"] = source.name
        row["ruta_archivo"] = str(source.resolve())
        return row

    if not rename:
        row["archivo_renombrado"] = f"{desired_stem}.pdf"
        row["ruta_archivo"] = str(source.resolve())
        return row

    try:
        target = unique_target_path(source, desired_stem)
        if target != source:
            source.rename(target)
        row["archivo_renombrado"] = target.name
        row["ruta_archivo"] = str(target.resolve())
        return row
    except Exception as exc:
        row["estado"] = "ERROR"
        row["observacion"] = f"Error renombrando: {exc}"
        row["archivo_renombrado"] = source.name
        row["ruta_archivo"] = str(source.resolve())
        return row


def process_pdfs_streaming(
    pdf_iter: Iterable[Path],
    rename: bool,
    writer: ExcelPartWriter,
    workers: int,
    progress_every: int,
    total_expected: int,
) -> Tuple[int, int, int, int]:
    total = 0
    ok = 0
    sin_numero = 0
    errores = 0

    # Si hay pocos PDFs, fuerza progreso en cada uno para que la GUI muestre movimiento.
    effective_every = 1 if total_expected <= 50 else progress_every

    def _emit_progress() -> None:
        # Emite progreso en el primer PDF, cada N segun configuracion y al final.
        if total == 1 or total % effective_every == 0 or total == total_expected:
            remaining = max(total_expected - total, 0)
            print(f"Progreso: {total} de {total_expected} | Faltan: {remaining}")

    if workers <= 1:
        for pdf in pdf_iter:
            row = extract_row_data(pdf)
            row = apply_rename(row, rename=rename)
            writer.append(row)

            total += 1
            if row.get("estado") == "OK":
                ok += 1
            elif row.get("estado") == "SIN_NUMERO":
                sin_numero += 1
            else:
                errores += 1

            _emit_progress()
        return total, ok, sin_numero, errores

    max_inflight = max(workers * 4, 16)
    inflight = set()

    with ProcessPoolExecutor(max_workers=workers) as executor:
        for pdf in pdf_iter:
            inflight.add(executor.submit(extract_row_data, pdf))

            while len(inflight) >= max_inflight:
                done, inflight = wait(inflight, return_when=FIRST_COMPLETED)
                for future in done:
                    row = future.result()
                    row = apply_rename(row, rename=rename)
                    writer.append(row)

                    total += 1
                    if row.get("estado") == "OK":
                        ok += 1
                    elif row.get("estado") == "SIN_NUMERO":
                        sin_numero += 1
                    else:
                        errores += 1

                    _emit_progress()

        while inflight:
            done, inflight = wait(inflight, return_when=FIRST_COMPLETED)
            for future in done:
                row = future.result()
                row = apply_rename(row, rename=rename)
                writer.append(row)

                total += 1
                if row.get("estado") == "OK":
                    ok += 1
                elif row.get("estado") == "SIN_NUMERO":
                    sin_numero += 1
                else:
                    errores += 1

                _emit_progress()

    return total, ok, sin_numero, errores


# =====================================================================
# NUEVO FORMATO HC (PDF NUEVOS)
# =====================================================================

# Encabezados de seccion conocidos que pueden aparecer como "tipo de servicio"
# entre la linea "Ingreso No:..." y la linea "HH:MM:SSDD/MM/YYYY ..."
_TIPOS_SERVICIO_HINTS = (
    "EVOLUCION", "EVOLUCIÓN",
    "NOTA", "CONSULTA", "VALORACION", "VALORACIÓN",
    "CONTROL", "INFORME", "ATENCION", "ATENCIÓN",
    "TRIAGE", "PSICOLOGIA", "PSICOLOGÍA", "ENFERMERIA", "ENFERMERÍA",
    "MEDICINA", "PSIQUIATRIA", "PSIQUIATRÍA", "TERAPIA",
)


def _extract_paciente_y_numero_ide(text: str) -> Tuple[str, str, str]:
    """Devuelve (paciente, numero_ide, tipo_id) desde la cabecera del PDF nuevo.

    Busca lineas tipo: " MAMIAN MOTTA GUSTAVO 12225097 CC 16/05/1953"
    """
    pattern = re.compile(
        r"^\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s\.\-']{2,80}?)\s+([A-Z0-9][A-Z0-9\-\.]{4,})\s+([A-Z]{2,4})\s+\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}",
        re.MULTILINE,
    )
    match = pattern.search(text)
    if match:
        return clean_value(match.group(1)), clean_value(match.group(2)), clean_value(match.group(3))
    return "", "", ""


def _extract_ingreso(text: str) -> str:
    match = re.search(r"Ingreso\s*No\s*[:\-]?\s*([A-Z0-9\-]+)", text, re.IGNORECASE)
    return clean_value(match.group(1)) if match else ""


def _extract_fecha_atencion_nuevo(text: str) -> str:
    """Busca el patron HH:MM:SSDD/MM/YYYY (sin espacio) o con espacio."""
    patterns = [
        r"\b\d{1,2}:\d{2}:\d{2}\s*([0-3]?\d[\/\-][0-1]?\d[\/\-]\d{2,4})\b",
        r"\bFecha\s*[:\-]?\s*([0-3]?\d[\/\-][0-1]?\d[\/\-]\d{2,4})",
    ]
    for p in patterns:
        m = re.search(p, text)
        if m:
            return clean_value(m.group(1))
    return ""


def _extract_tipo_servicio(text: str) -> str:
    """Extrae el titulo del servicio que aparece entre 'Ingreso No:' y la linea de fecha."""
    # Buscar todo lo que va despues de la primera ocurrencia de "Ingreso No:..."
    m = re.search(r"Ingreso\s*No\s*[:\-]?\s*[A-Z0-9\-]+\s*\n(.+?)(?:\n\s*\d{1,2}:\d{2}:\d{2}|\n\s*Fecha\b)",
                  text, re.IGNORECASE | re.DOTALL)
    if not m:
        return ""

    block = m.group(1)
    # Tomar lineas no vacias y elegir la primera que contenga una palabra clave conocida.
    lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
    for ln in lines:
        upper = ln.upper()
        if any(h in upper for h in _TIPOS_SERVICIO_HINTS):
            # Limpiar puntos finales
            return clean_value(ln.rstrip(". "))
    # Fallback: primera linea no vacia
    return clean_value(lines[0]) if lines else ""


def _extract_diagnostico(text: str) -> Tuple[str, str, str, str]:
    """Devuelve (tipo, clase, codigo, descripcion) del primer diagnostico de la tabla.

    Tablas tipo:
        Tipo Clase Diagnostico Observaciones
        Diagnosticos
        Principal Confirmado Repetido F319 TRASTORNO AFECTIVO BIPOLAR, NO ESPECIFICADO
        Principal Impresion Diagnostica F319 TRASTORNO AFECTIVO BIPOLAR, NO ESPECIFICADO
    """
    # Localizar bloque de diagnosticos
    m_block = re.search(
        r"(?:IMPRESI[OÓ]N\s+DIAGN[OÓ]STICA|DIAGNOSTICOS|DIAGN[OÓ]STICOS)\b(.+?)(?:\n\s*(?:\d+\s*\.\s*)?(?:PLAN|REMISION|REMISIÓN|ORDENES|ÓRDENES|NOVEDADES)\b|\Z)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    block = m_block.group(1) if m_block else text

    # Patron clave: codigo CIE10 (letra + 3-4 digitos opcional con letra) seguido de descripcion
    line_pat = re.compile(
        r"(Principal|Relacionado|Secundario)\s+([A-Za-zÁÉÍÓÚáéíóúÑñ\s]+?)\s+([A-Z]\d{2,4}[A-Z0-9]?)\s+([^\n\r]+)",
    )
    m = line_pat.search(block)
    if m:
        tipo = clean_value(m.group(1))
        clase = clean_value(m.group(2))
        codigo = clean_value(m.group(3))
        descripcion = clean_value(m.group(4))
        # Quitar columna "Observaciones" residual si quedo en blanco
        descripcion = re.sub(r"\s{2,}.*$", "", descripcion).strip()
        return tipo, clase, codigo, descripcion
    return "", "", "", ""


def _extract_nota(text: str) -> str:
    """Extrae el campo Nota.

    - Para HC tipo enfermeria: lineas inmediatamente despues de la etiqueta 'Nota'
      hasta el siguiente campo (linea que empieza con palabra capitalizada o bloque).
    - Para HC psicologia: si no hay 'Nota', cae a 'Motivo consulta'.
    """
    # Caso 1: bloque "Nota ... <texto>"
    m = re.search(
        r"\n([^\n]{20,}?)\s*Nota\s*\n(.+?)(?=\n[A-ZÁÉÍÓÚÑ][a-zñáéíóú]+\s|\nEXAMEN|\nDIAGNOSTICO|\Z)",
        text,
        re.DOTALL,
    )
    # El primer grupo seria texto previo, no util. Probemos uno mas directo.
    m2 = re.search(
        r"(?:^|\n)\s*([^\n]{30,})\s*Nota\s*$",
        text,
        re.MULTILINE,
    )
    if m2:
        # En el formato extraido, "Nota" aparece al final de una linea cuyo inicio
        # contiene el primer renglon de la nota; las lineas siguientes son la continuacion.
        idx = m2.end()
        first_line = m2.group(1).strip()
        rest = text[idx:idx + 4000]
        # Cortar al siguiente "<Algo>" en titulo (linea con etiqueta sin contenido) o EXAMEN/DIAGNOST.
        cut = re.search(r"\n\s*(EXAMEN|DIAGNOSTIC|G[ÉE]NERO|Otro\s*G|S[ií]ntom|ORDENES|ORDENES\b|11\.|10\.|9\.)", rest)
        block = rest[: cut.start()] if cut else rest
        # Limpiar lineas vacias dobles
        joined = " ".join(line.strip() for line in (first_line + " " + block).splitlines() if line.strip())
        return clean_value(joined)[:3000]

    # Caso 2: psicologia -> Motivo consulta
    m3 = re.search(
        r"Motivo\s+consulta\s*\n?(.+?)(?:\nFecha\s+semana|\n\d+\.\s|\Z)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    if m3:
        joined = " ".join(line.strip() for line in m3.group(1).splitlines() if line.strip())
        return clean_value(joined)[:3000]

    return ""


def _extract_novedades(text: str) -> Tuple[str, str]:
    """Extrae el contenido de la seccion '12. NOVEDADES'.

    Devuelve (novedades, fecha_novedad). Busca primero el bloque
    delimitado por '12. NOVEDADES' y la siguiente seccion numerada o
    final del documento. Tiene fallback si no aparece la numeracion.
    """
    # Bloque desde "12. NOVEDADES" hasta la siguiente seccion (13.) o fin.
    m = re.search(
        r"(?is)\b12\.\s*NOVEDADES\b(.+?)(?=\n\s*1[3-9]\.\s|\n\s*\d{2,}\.\s|\Z)",
        text,
    )
    if not m:
        # Fallback: titulo NOVEDADES sin numero.
        m = re.search(
            r"(?is)\bNOVEDADES\b\s*\n(.+?)(?=\n\s*\d+\.\s|\Z)",
            text,
        )
    if not m:
        return "", ""

    block = m.group(1)

    # Campo "Novedades  <texto>"  (texto puede continuar en lineas siguientes)
    novedades = ""
    n_match = re.search(
        r"(?im)^\s*Novedades\s*[:\-]?\s*(.+?)(?=\n\s*Fecha\s+de\s+novedad\b|\n\s*\d+\.\s|\Z)",
        block,
    )
    if n_match:
        novedades = clean_value(
            " ".join(ln.strip() for ln in n_match.group(1).splitlines() if ln.strip())
        )[:1000]

    # Campo "Fecha de novedad  <fecha>"
    fecha = ""
    f_match = re.search(
        r"(?im)^\s*Fecha\s+de\s+novedad\s*[:\-]?\s*(.+?)\s*$",
        block,
    )
    if f_match:
        fecha = clean_value(f_match.group(1))

    return novedades, fecha


def detect_fields_nuevo(text: str) -> Dict[str, str]:
    paciente, numero_ide, _tipo_id = _extract_paciente_y_numero_ide(text)
    diag_tipo, diag_clase, diag_codigo, diag_desc = _extract_diagnostico(text)
    novedades, fecha_novedad = _extract_novedades(text)

    return {
        "paciente": paciente,
        "numero_ide": numero_ide,
        "ingreso": _extract_ingreso(text),
        "tipo_servicio": _extract_tipo_servicio(text),
        "fecha_atencion": _extract_fecha_atencion_nuevo(text),
        "diagnostico_tipo": diag_tipo,
        "diagnostico_clase": diag_clase,
        "diagnostico_codigo": diag_codigo,
        "diagnostico_descripcion": diag_desc,
        "nota": _extract_nota(text),
        "novedades": novedades,
        "fecha_novedad": fecha_novedad,
    }


def extract_row_nuevo(pdf_path: Path) -> Dict[str, str]:
    row: Dict[str, str] = {
        "ruta_archivo": str(pdf_path.resolve()),
        "archivo": pdf_path.name,
        "paciente": "",
        "numero_ide": "",
        "ingreso": "",
        "tipo_servicio": "",
        "fecha_atencion": "",
        "diagnostico_tipo": "",
        "diagnostico_clase": "",
        "diagnostico_codigo": "",
        "diagnostico_descripcion": "",
        "nota": "",
        "novedades": "",
        "fecha_novedad": "",
        "estado": "OK",
        "observacion": "",
    }
    try:
        text = extract_pdf_text(pdf_path)
        row.update(detect_fields_nuevo(text))
        if not row["numero_ide"] and not row["paciente"]:
            row["estado"] = "SIN_DATOS"
            row["observacion"] = "No se detectaron datos basicos del paciente"
        return row
    except Exception as exc:
        row["estado"] = "ERROR"
        row["observacion"] = str(exc)
        return row


def process_pdfs_nuevos_streaming(
    pdf_iter: Iterable[Path],
    writer: ExcelPartWriter,
    workers: int,
    progress_every: int,
    total_expected: int,
) -> Tuple[int, int, int]:
    total = 0
    ok = 0
    errores = 0

    effective_every = 1 if total_expected <= 50 else progress_every

    def _emit_progress() -> None:
        if total == 1 or total % effective_every == 0 or total == total_expected:
            remaining = max(total_expected - total, 0)
            print(f"Progreso: {total} de {total_expected} | Faltan: {remaining}")

    if workers <= 1:
        for pdf in pdf_iter:
            row = extract_row_nuevo(pdf)
            writer.append(row)
            total += 1
            if row.get("estado") == "OK":
                ok += 1
            else:
                errores += 1
            _emit_progress()
        return total, ok, errores

    max_inflight = max(workers * 4, 16)
    inflight = set()
    with ProcessPoolExecutor(max_workers=workers) as executor:
        for pdf in pdf_iter:
            inflight.add(executor.submit(extract_row_nuevo, pdf))
            while len(inflight) >= max_inflight:
                done, inflight = wait(inflight, return_when=FIRST_COMPLETED)
                for fut in done:
                    row = fut.result()
                    writer.append(row)
                    total += 1
                    if row.get("estado") == "OK":
                        ok += 1
                    else:
                        errores += 1
                    _emit_progress()
        while inflight:
            done, inflight = wait(inflight, return_when=FIRST_COMPLETED)
            for fut in done:
                row = fut.result()
                writer.append(row)
                total += 1
                if row.get("estado") == "OK":
                    ok += 1
                else:
                    errores += 1
                _emit_progress()

    return total, ok, errores


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Procesa PDF para informe/renombre, o encarpetado por numero_documento desde un Excel.")
    parser.add_argument(
        "--accion",
        choices=["procesar-pdfs", "encarpetar-desde-excel", "informe-nuevos"],
        default="procesar-pdfs",
        help="Accion a ejecutar",
    )
    parser.add_argument("--entrada", help="Carpeta que contiene los PDF. Si se omite, se solicita por consola")
    parser.add_argument("--salida", default="INFORME_PDFS.xlsx", help="Ruta del Excel de salida")
    parser.add_argument("--sin-renombrar", action="store_true", help="Solo genera informe sin renombrar PDF")
    parser.add_argument("--recursivo", action="store_true", help="Busca PDF tambien en subcarpetas")
    parser.add_argument("--workers", type=int, default=max((os.cpu_count() or 2) - 1, 1), help="Cantidad de procesos en paralelo")
    parser.add_argument("--max-filas-excel", type=int, default=1_000_000, help="Maximo de filas por archivo Excel")
    parser.add_argument("--progreso-cada", type=int, default=1000, help="Imprime progreso cada N PDF")
    parser.add_argument("--excel-entrada", help="Excel con columnas ruta_archivo y numero de documento para encarpetar")
    parser.add_argument("--excel-salida-encarpetado", default="INFORME_ENCARPETADO.xlsx", help="Ruta del informe de encarpetado")
    parser.add_argument("--col-ruta", default="ruta_archivo", help="Nombre de la columna de ruta del archivo")
    parser.add_argument("--col-documento", default="numero de documento", help="Nombre de la columna de numero de documento")

    args = parser.parse_args()

    if args.accion == "encarpetar-desde-excel":
        if not args.excel_entrada:
            raise SystemExit("Debes indicar --excel-entrada para la accion encarpetar-desde-excel")

        excel_input = Path(args.excel_entrada).expanduser().resolve()
        if not excel_input.exists() or not excel_input.is_file():
            raise SystemExit(f"No existe el Excel de entrada: {excel_input}")

        excel_output = Path(args.excel_salida_encarpetado).expanduser().resolve()

        total, ok, errores = process_excel_foldering(
            excel_input=excel_input,
            excel_output=excel_output,
            ruta_column_name=args.col_ruta,
            documento_column_name=args.col_documento,
        )

        print(f"Proceso de encarpetado finalizado. Total: {total} | OK: {ok} | ERROR: {errores}")
        print(f"Informe de encarpetado: {excel_output}")
        return

    if args.accion == "informe-nuevos":
        entrada = args.entrada
        if not entrada:
            entrada = input("Ingrese la ruta de la carpeta con PDF NUEVOS: ").strip().strip('"')
        if not entrada:
            raise SystemExit("Debes indicar --entrada con la carpeta de PDF NUEVOS")

        input_dir = resolve_input_dir(entrada)
        output_excel = Path(args.salida).expanduser().resolve()

        total_detectado = count_pdf_files(input_dir, args.recursivo)
        if total_detectado == 0:
            raise SystemExit("No se encontraron archivos PDF en la carpeta indicada")

        print(f"Total detectado: {total_detectado}")

        writer = ExcelPartWriter(
            output_excel,
            max_rows_per_file=max(args.max_filas_excel, 1),
            headers=HEADERS_NUEVOS,
            sheet_name="Informe_Nuevos",
        )
        pdf_iter = iter_pdf_files(input_dir, args.recursivo)

        total, ok, errores = process_pdfs_nuevos_streaming(
            pdf_iter=pdf_iter,
            writer=writer,
            workers=max(args.workers, 1),
            progress_every=max(args.progreso_cada, 1),
            total_expected=total_detectado,
        )
        generated = writer.close()
        print(f"Proceso finalizado. Total: {total} | OK: {ok} | ERROR: {errores}")
        if len(generated) == 1:
            print(f"Informe generado en: {generated[0]}")
        else:
            print("Informes generados:")
            for p in generated:
                print(f" - {p}")
        return

    entrada = args.entrada
    if not entrada:
        entrada = input("Ingrese la ruta de la carpeta con PDF: ").strip().strip('"')
    if not entrada:
        raise SystemExit("Debes indicar una ruta de carpeta en --entrada o escribirla en consola")

    input_dir = resolve_input_dir(entrada)
    output_excel = Path(args.salida).expanduser().resolve()

    rename = not args.sin_renombrar

    total_detectado = count_pdf_files(input_dir, args.recursivo)
    if total_detectado == 0:
        raise SystemExit("No se encontraron archivos PDF en la carpeta indicada")

    print(f"Total detectado: {total_detectado}")

    writer = ExcelPartWriter(output_excel, max_rows_per_file=max(args.max_filas_excel, 1))
    pdf_iter = iter_pdf_files(input_dir, args.recursivo)

    total, ok, sin_numero, errores = process_pdfs_streaming(
        pdf_iter=pdf_iter,
        rename=rename,
        writer=writer,
        workers=max(args.workers, 1),
        progress_every=max(args.progreso_cada, 1),
        total_expected=total_detectado,
    )

    generated_reports = writer.close()

    print(f"Proceso finalizado. Total: {total} | OK: {ok} | SIN_NUMERO: {sin_numero} | ERROR: {errores}")
    if len(generated_reports) == 1:
        print(f"Informe generado en: {generated_reports[0]}")
    else:
        print("Informes generados:")
        for path in generated_reports:
            print(f" - {path}")


if __name__ == "__main__":
    main()
